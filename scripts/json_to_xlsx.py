#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl", "requests", "beautifulsoup4", "tabulate"]
# ///
"""Professional Excel export for US Small-Cap DD Scanner outputs."""
from __future__ import annotations

import argparse, importlib.util, json, os, sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

WORKSPACE = Path(os.path.expanduser("~/.openclaw/workspace"))
SKILL_DIR = WORKSPACE / "skills" / "us-smallcap-dd"
RESULTS_DIR = WORKSPACE / "data" / "dd_results"
WATCHLIST = Path(os.path.expanduser("~/.openclaw/watchlist.json"))

NAVY="0A1929"; GREEN="2A9D8F"; ORANGE="F4A261"; RED="E63946"; GREY="5A6C7D"; LIGHT="F0F4F8"; WHITE="FFFFFF"; BORDER="D0D0D0"
HEADER_FILL=PatternFill("solid", fgColor=NAVY); TITLE_FILL=PatternFill("solid", fgColor=NAVY)
GREEN_FILL=PatternFill("solid", fgColor=GREEN); ORANGE_FILL=PatternFill("solid", fgColor=ORANGE); RED_FILL=PatternFill("solid", fgColor=RED)
ZEBRA_FILL=PatternFill("solid", fgColor=LIGHT); WHITE_FILL=PatternFill("solid", fgColor=WHITE)
HEADER_FONT=Font("Calibri", 11, bold=True, color=WHITE); TITLE_FONT=Font("Calibri", 14, bold=True, color=WHITE)
SUBTITLE_FONT=Font("Calibri", 10, italic=True, color=GREY); NORMAL_FONT=Font("Calibri", 10); BOLD_FONT=Font("Calibri", 10, bold=True)
VERDICT_FONT=Font("Calibri", 10, bold=True, color=WHITE)
THIN_BORDER=Border(left=Side("thin", color=BORDER), right=Side("thin", color=BORDER), top=Side("thin", color=BORDER), bottom=Side("thin", color=BORDER))

LAYER_KEYS=[("business_quality","Business (Phase 1)"),("financial_performance","Financial (Phase 1)"),("cash_flow_reality","Cash Flow"),("debt_capital_structure","Debt/Capital"),("ownership_governance","Governance"),("valuation","Valuation")]


def num(x: Any) -> float | None:
    if x in (None, "", "N/A", "None", "-"): return None
    try:
        if isinstance(x, str):
            s=x.strip().replace(",", "").replace("$", "")
            pct=s.endswith("%");
            if pct: s=s[:-1]
            mult=1.0
            if s.upper().endswith("T"): mult=1e12; s=s[:-1]
            elif s.upper().endswith("B"): mult=1e9; s=s[:-1]
            elif s.upper().endswith("M"): mult=1e6; s=s[:-1]
            elif s.upper().endswith("K"): mult=1e3; s=s[:-1]
            v=float(s)*mult
            return v/100 if pct else v
        return float(x)
    except Exception:
        return None


def market_cap_value(row: dict[str, Any]) -> float | None:
    v = row.get("market_cap") or row.get("marketCap") or row.get("key_metrics",{}).get("market_cap")
    mv = num(v)
    if mv is not None: return mv
    m = num(row.get("market_cap_M"))
    return m*1e6 if m is not None else None


def fmt_mcap(v: Any) -> str:
    x=num(v)
    if x is None: return "N/A"
    if x >= 1e12: return f"${x/1e12:.1f}T"
    if x >= 1e9: return f"${x/1e9:.1f}B"
    if x >= 1e6: return f"${x/1e6:.0f}M"
    return f"${x:,.0f}"


def pct_value(v: Any) -> float | None:
    x=num(v)
    if x is None: return None
    return x/100 if abs(x)>1.5 else x


def score_value(row: dict[str, Any]) -> float | None:
    s = row.get("composite_score", row.get("score"))
    x=num(s)
    if x is not None:
        return round(x*10,1) if 0 <= x <= 1 else round(x,1)
    scores = row.get("layer_scores") or row.get("dd_scores") or {}
    vals=[num(scores.get(k)) for k,_ in LAYER_KEYS if num(scores.get(k)) is not None]
    return round(sum(vals)/len(vals),1) if vals else None


def verdict_fill(verdict: Any) -> PatternFill:
    v=str(verdict or "").upper().replace(" ", "_")
    if "DEEP_VALUE" in v or ("DEEP" in v and "VALUE" in v): return GREEN_FILL
    if "VALUE_TRAP" in v or "TRAP" in v: return ORANGE_FILL
    if "NOT_INVEST" in v or "NOT" in v or "AVOID" in v: return RED_FILL
    return PatternFill("solid", fgColor=GREY)


def normalize_results(data: Any) -> list[dict[str, Any]]:
    if isinstance(data, list): return [r for r in data if isinstance(r, dict)]
    if not isinstance(data, dict): return []
    if isinstance(data.get("results"), list): return data["results"]
    if isinstance(data.get("top_picks"), list): return data["top_picks"]
    if data.get("ticker"): return [data]
    return []


def apply_title(ws, title: str, subtitle: str, last_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.cell(1,1,title).fill=TITLE_FILL; ws.cell(1,1).font=TITLE_FONT; ws.cell(1,1).alignment=Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws.cell(2,1,subtitle).font=SUBTITLE_FONT
    ws.row_dimensions[1].height=24


def style_headers(ws, row: int, headers: list[str]) -> None:
    for c,h in enumerate(headers,1):
        cell=ws.cell(row,c,h); cell.fill=HEADER_FILL; cell.font=HEADER_FONT; cell.border=THIN_BORDER; cell.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.auto_filter.ref=f"A{row}:{get_column_letter(len(headers))}{ws.max_row}"
    ws.freeze_panes="A5"


def score_style(cell) -> None:
    v=num(cell.value)
    if v is None: return
    cell.number_format='0.0'
    if v > 7: cell.fill=GREEN_FILL; cell.font=Font("Calibri",10,bold=True,color=WHITE)
    elif v >= 4: cell.fill=ORANGE_FILL; cell.font=Font("Calibri",10,bold=True,color=WHITE)
    else: cell.fill=RED_FILL; cell.font=Font("Calibri",10,bold=True,color=WHITE)


def finish_sheet(ws, header_row: int, max_width: int = 30) -> None:
    # Set column widths based on header text + data
    for col in range(1, ws.max_column + 1):
        header_text = str(ws.cell(header_row, col).value or "")
        # Min width from header, max from content scan
        max_len = max(len(header_text), 6)
        for row in range(header_row + 1, ws.max_row + 1):
            val = ws.cell(row, col).value
            if val is not None:
                max_len = max(max_len, min(max_width, len(str(val)) + 2))
        ws.column_dimensions[get_column_letter(col)].width = max(6, min(max_width, max_len))
    # Zebra stripe + borders + vertical alignment
    for r in range(header_row + 1, ws.max_row + 1):
        base = ZEBRA_FILL if (r - header_row) % 2 == 0 else WHITE_FILL
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.border = THIN_BORDER
            if not cell.font or cell.font == Font():
                cell.font = NORMAL_FONT
            if cell.fill.fill_type is None:
                cell.fill = base
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_score_conditional(ws, col: int, first: int, last: int) -> None:
    if last < first: return
    rng=f"{get_column_letter(col)}{first}:{get_column_letter(col)}{last}"
    ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThan', formula=['7'], fill=GREEN_FILL, font=Font(color=WHITE,bold=True)))
    ws.conditional_formatting.add(rng, CellIsRule(operator='between', formula=['4','7'], fill=ORANGE_FILL, font=Font(color=WHITE,bold=True)))
    ws.conditional_formatting.add(rng, CellIsRule(operator='lessThan', formula=['4'], fill=RED_FILL, font=Font(color=WHITE,bold=True)))


def convert_screener(data: dict[str, Any], output_path: Path):
    rows=sorted(normalize_results(data), key=lambda r: score_value(r) or -999, reverse=True)
    wb=Workbook(); ws=wb.active; ws.title="Summary"; ws.sheet_properties.tabColor="1D9BF0"
    headers=["Rank","Ticker","Sector","Industry","Market Cap","Composite Score","P/E","P/B","P/FCF","FCF Yield","Debt/Eq","ROE","Revenue Growth","Verdict","Next Step"]
    passing=data.get('candidates_passing_filters', len(rows)); screened=data.get('candidates_screened', len(rows)); date=data.get('date', datetime.now(timezone.utc).date().isoformat())
    apply_title(ws, f"Russell 2000 Value Screener — {date}", f"Screened: {screened} tickers | Passing: {passing} | Phase 1 only (run dd_scan.py --ticker TICKER for full DD) | Source: FMP /stable/ + Polygon + UW + SEC EDGAR", len(headers))
    style_headers(ws,4,headers)
    DASH = "—"
    has_any_dd = any((r.get('dd_verdict') or r.get('verdict')) for r in rows)
    for i,r in enumerate(rows,5):
        km=r.get('key_metrics',{}) if isinstance(r.get('key_metrics'),dict) else {}
        verdict_raw = r.get('dd_verdict') or r.get('verdict')
        verdict_display = verdict_raw or "Pending DD"
        summary = r.get('dd_summary') or r.get('one_line_summary') or r.get('summary')
        next_step = summary if summary else f"uv run scripts/dd_scan.py --ticker {r.get('ticker')}"
        vals=[i-4,r.get('ticker'),r.get('sector') or km.get('sector') or DASH,r.get('industry') or km.get('industry') or DASH,fmt_mcap(market_cap_value(r)),score_value(r),num(r.get('pe') or km.get('p_e')),num(r.get('pb') or km.get('p_b')),num(r.get('pfcf') or km.get('p_fcf')),pct_value(r.get('fcf_yield') if r.get('fcf_yield') is not None else km.get('fcf_margin_latest')),num(r.get('debt_equity') or km.get('debt_to_equity_finviz')),pct_value(r.get('roe') or km.get('roe')),pct_value(r.get('revenue_growth') if r.get('revenue_growth') is not None else km.get('revenue_cagr_5y')),verdict_display,next_step]
        for c,v in enumerate(vals,1):
            cell = ws.cell(i,c,v if v is not None else DASH)
            cell.border=THIN_BORDER
        # Number formats: P/E, P/B, P/FCF, Debt/Eq are plain decimal numbers
        for c in [7,8,9,11]:
            if isinstance(ws.cell(i,c).value, (int, float)):
                ws.cell(i,c).number_format='0.00'
        # Percentages: FCF Yield, ROE, Revenue Growth
        for c in [10,12,13]:
            if isinstance(ws.cell(i,c).value, (int, float)):
                ws.cell(i,c).number_format='0.0%'
        score_style(ws.cell(i,6))
        ws.cell(i,14).fill=verdict_fill(verdict_raw if verdict_raw else None)
        ws.cell(i,14).font=VERDICT_FONT
        ws.cell(i,14).alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)
    add_score_conditional(ws,6,5,ws.max_row); finish_sheet(ws,4, max_width=42)
    # Note row at the bottom
    if not has_any_dd:
        note_row = ws.max_row + 2
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(headers))
        nc = ws.cell(note_row, 1, "Note: Screener (Phase 1) only — no full DD verdict/layer scores. Run: uv run scripts/dd_scan.py --ticker TICKER for the 6-layer DD.")
        nc.font = Font("Calibri", 10, italic=True, color=GREY)
        nc.alignment = Alignment(horizontal='left', vertical='center')
        nc.fill = PatternFill("solid", fgColor=LIGHT)

    ws2=wb.create_sheet("Details"); headers2=["Ticker","Verdict","Composite Score"]+[h for _,h in LAYER_KEYS]+["Red Flags","Catalysts","Next Step"]
    apply_title(ws2, f"Russell 2000 Value Screener Details — {date}", f"Screened: {screened} tickers | Passing: {passing} | Phase 1 only — full DD requires dd_scan.py", len(headers2))
    style_headers(ws2,4,headers2)
    for i,r in enumerate(rows,5):
        scores=r.get('layer_scores') or r.get('dd_scores') or {}
        verdict_raw = r.get('dd_verdict') or r.get('verdict')
        verdict_display = verdict_raw or "Pending DD"
        red_flags_list = r.get('red_flags') or []
        catalysts_list = r.get('catalysts') or []
        red_flags_text = ', '.join(map(str, red_flags_list)) if red_flags_list else DASH
        catalysts_text = ', '.join(map(str, catalysts_list)) if catalysts_list else DASH
        next_step = f"uv run scripts/dd_scan.py --ticker {r.get('ticker')}" if not verdict_raw else (r.get('one_line_summary') or r.get('dd_summary') or r.get('summary') or DASH)
        layer_vals = [num(scores.get(k)) for k,_ in LAYER_KEYS]
        layer_vals = [(v if v is not None else "Pending DD") for v in layer_vals]
        score_v = score_value(r)
        score_display = score_v if score_v is not None else "Pending DD"
        vals=[r.get('ticker'), verdict_display, score_display] + layer_vals + [red_flags_text, catalysts_text, next_step]
        for c,v in enumerate(vals,1): ws2.cell(i,c,v).border=THIN_BORDER
        ws2.cell(i,2).fill=verdict_fill(verdict_raw if verdict_raw else None)
        ws2.cell(i,2).font=VERDICT_FONT
        ws2.cell(i,2).alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)
        for c in range(3,10):
            if isinstance(ws2.cell(i,c).value, (int, float)):
                score_style(ws2.cell(i,c))
            else:
                ws2.cell(i,c).font=Font("Calibri", 9, italic=True, color=GREY)
                ws2.cell(i,c).alignment=Alignment(horizontal='center', vertical='center')
    for c in range(3,10): add_score_conditional(ws2,c,5,ws2.max_row)
    finish_sheet(ws2,4, max_width=50)
    wb.save(output_path); return output_path


def convert_dd(data: dict[str, Any], output_path: Path, title_date: str | None = None):
    rows=sorted(normalize_results(data), key=lambda r: score_value(r) or -999, reverse=True)
    wb=Workbook(); ws=wb.active; ws.title="DD Results"; date=title_date or data.get('date') or datetime.now(timezone.utc).date().isoformat()
    headers=["Ticker","Verdict","Composite Score"]+[h for _,h in LAYER_KEYS]+["1-Line Summary"]
    apply_title(ws, f"Due Diligence Scan — {date}", f"Screened: {len(rows)} tickers | Passing: {sum(1 for r in rows if 'DEEP' in str(r.get('verdict') or r.get('dd_verdict')).upper())} | Source: FMP /stable/ + Polygon + UW + SEC EDGAR", len(headers))
    style_headers(ws,4,headers)
    for i,r in enumerate(rows,5):
        scores=r.get('layer_scores') or r.get('dd_scores') or {}; vals=[r.get('ticker'),r.get('verdict') or r.get('dd_verdict') or '',score_value(r)] + [num(scores.get(k)) for k,_ in LAYER_KEYS] + [r.get('one_line_summary') or r.get('dd_summary') or r.get('summary') or '']
        for c,v in enumerate(vals,1): ws.cell(i,c,v).border=THIN_BORDER
        ws.cell(i,2).fill=verdict_fill(ws.cell(i,2).value); ws.cell(i,2).font=VERDICT_FONT
        for c in range(3,10): score_style(ws.cell(i,c))
    for c in range(3,10): add_score_conditional(ws,c,5,ws.max_row)
    finish_sheet(ws,4); wb.save(output_path); return output_path


def load_watchlist_tickers() -> list[str]:
    data=json.loads(WATCHLIST.read_text())
    if isinstance(data, dict) and isinstance(data.get('tickers'), list):
        data = data['tickers']
    found=[]
    def walk(x):
        if isinstance(x, str):
            s=x.upper().strip()
            if s and s.replace('.','').replace('-','').isalnum() and len(s)<=8: found.append(s)
        elif isinstance(x, dict):
            if 'ticker' in x: walk(x['ticker'])
            elif 'symbol' in x: walk(x['symbol'])
            else:
                for v in x.values(): walk(v)
        elif isinstance(x, list):
            for v in x: walk(v)
    walk(data)
    return sorted(dict.fromkeys(found))


def run_watchlist() -> dict[str, Any]:
    spec=importlib.util.spec_from_file_location('dd_scan', SKILL_DIR/'scripts'/'dd_scan.py')
    mod=importlib.util.module_from_spec(spec); assert spec and spec.loader; sys.modules['dd_scan']=mod; spec.loader.exec_module(mod)  # type: ignore
    tickers=load_watchlist_tickers(); results=[]
    for t in tickers:
        print(f"DD {t}...")
        try: results.append(mod.analyze_ticker(t))
        except Exception as exc: results.append({'ticker':t,'verdict':'NOT_INVESTABLE','score':0,'summary':f'ERROR: {exc}','red_flags':[str(exc)]})
    return {'date': datetime.now(timezone.utc).date().isoformat(), 'generated_at': datetime.now(timezone.utc).isoformat(), 'tickers': tickers, 'results': results}


def main() -> int:
    p=argparse.ArgumentParser(description='Convert JSON DD results to professional Excel')
    p.add_argument('--date'); p.add_argument('--dd', action='store_true'); p.add_argument('--all', action='store_true'); p.add_argument('--watchlist', action='store_true')
    args=p.parse_args(); RESULTS_DIR.mkdir(parents=True, exist_ok=True); target=args.date or datetime.now(timezone.utc).date().isoformat()
    if args.watchlist:
        data=run_watchlist(); json_path=RESULTS_DIR/f"{data['date']}_watchlist_dd.json"; json_path.write_text(json.dumps(data, indent=2, default=str)); out=json_path.with_suffix('.xlsx'); convert_dd(data,out,data['date']); print(f"Saved: {out}"); return 0
    if args.all:
        for f in sorted(RESULTS_DIR.glob('*.json')):
            data=json.loads(f.read_text()); out=f.with_suffix('.xlsx'); (convert_screener if 'screener' in f.name else convert_dd)(data,out); print(f"{f.name} → {out.name}")
        return 0
    if args.dd:
        cands=[RESULTS_DIR/f"{target}.json", *sorted(RESULTS_DIR.glob(f"{target}_*_dd.json"))]
        json_path=next((x for x in cands if x.exists()), None)
        if not json_path: print(f"No DD file found for {target}"); return 1
        data=json.loads(json_path.read_text()); out=json_path.with_suffix('.xlsx'); convert_dd(data,out,target); print(f"Saved: {out}"); return 0
    json_path=RESULTS_DIR/f"{target}_screener.json"
    if not json_path.exists(): print(f"No screener file found for {target}"); return 1
    data=json.loads(json_path.read_text()); out=json_path.with_suffix('.xlsx'); convert_screener(data,out); print(f"Saved: {out}"); return 0

if __name__ == '__main__': raise SystemExit(main())
